"""
SAP Task Handlers
=================
Mỗi task auto là 1 hàm riêng trong file này.
Hàm nhận (filepath, session) và trả về dict kết quả.

Khi thêm task mới:
1. Viết hàm mới ở đây
2. Tạo TaskConfig trên web, điền handler_module = "tasks.handlers.ten_ham"
"""
import os
import logging
import time
from django.utils import timezone
from .SAP_scripts.sap_base import SapGuiClient
from .SAP_scripts.tcode_ob08 import TCodeOB08
from .SAP_scripts.tcode_md12 import TCodeMD12
from .SAP_scripts.tcode_mm02 import TCodeMM02
from .SAP_scripts.tcode_me22 import TCodeME22
from .SAP_scripts.tcode_me12 import TCodeME12
from .SAP_scripts.tcode_mk01 import TCodeMK01
from .SAP_scripts.tcode_cs02 import TCodeCS02
from .SAP_scripts.tcode_vl32n import TCodeVL32N
from .SAP_scripts.tcode_me52 import TCodeME52
from .SAP_scripts.tcode_qa11 import TCodeQA11



log = logging.getLogger(__name__)


def default_handler(filepath, session=None, task=None):
    """Handler mặc định - chỉ log, không làm gì"""
    log.info(f"[DEFAULT] File: {filepath}")
    return {
        "status": "success",
        "message": "Default handler - no action taken",
        "rows": 0
    }


def exchange_rate(filepath, session=None, task=None):
    """
    OB08 - Auto input Exchange Rate into SAP
    
    Tham số 1 (task.param1): USD/JPY-B20-D20-1, USD/VND-B30-D30-1000
    Format: CURRENCY_PAIR-CELL_NAME-CELL_VALUE-DIVISOR, ...
    """
    log.info(f"[OB08] Bắt đầu xử lý: {os.path.basename(filepath)}")
    start = time.time()

    try:
        # ===== 1. Đọc file Excel =====
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        # ===== 2. Parse tham số 1 =====
        rows_data = []
        today = timezone.localtime(timezone.now()).strftime("%d.%m.%Y")

        if task and task.param1:
            items = [item.strip() for item in task.param1.split(',')]
            
            for item in items:
                parts = item.split('-')
                if len(parts) != 4:
                    log.warning(f"  Bỏ qua format không hợp lệ: {item}")
                    continue
                
                currency_pair = parts[0].strip()
                cell_name = parts[1].strip()
                cell_value = parts[2].strip()
                divisor = parts[3].strip()

                name_value = ws[cell_name].value
                rate_value = ws[cell_value].value
                
                if rate_value is None:
                    log.warning(f"  Ô {cell_value} không có giá trị")
                    continue

                try:
                    divisor_num = float(divisor)
                    if divisor_num != 0:
                        rate_value = f"{float(rate_value) / divisor_num:.3f}"
                except ValueError:
                    log.warning(f"  Divisor không hợp lệ: {divisor}")
                    continue

                if '/' in currency_pair:
                    currency_from, currency_to = currency_pair.split('/')
                else:
                    currency_from = currency_pair
                    currency_to = 'VND'
                
                rows_data.append({
                    'kurst': 'M',
                    'gdatu': today,
                    'kursm': str(rate_value),
                    'fcurr': currency_to,
                    'tcurr': currency_from,
                })
                
                log.info(f"  {currency_pair}: {cell_name}={name_value}, {cell_value}={rate_value}")

        if not rows_data:
            return {"status": "skipped", "message": "Không có dữ liệu hoặc param1 rỗng", "rows": 0}

        log.info(f"  Đọc được {len(rows_data)} tỷ giá")

        if not task or not task.sap_user:
            return {
                "status": "success",
                "message": f"Validated (no SAP user): {rows_data}",
                "rows": len(rows_data)
            }

        sap_client = task.sap_user.client
        sap_username = task.sap_user.username
        sap_password = task.sap_user.password
        log.info(f"  SAP User: {sap_client}/{sap_username}")

        time.sleep(2)

        with SapGuiClient(
            sap_entry_name='V2Q',
            client_no=sap_client,
            username=sap_username,
            password=sap_password,
        ) as sap:
            
            st = sap.last_login_status
            if sap.is_error(st) or ("name or password is incorrect" in (st.text or "").lower()):
                return {
                    "status": "error",
                    "message": f"SAP Login Error: {st.text}",
                    "rows": 0,
                    "duration": round(time.time() - start, 2)
                }
            
            log.info(f"  Đăng nhập SAP thành công")
            
            ob08 = TCodeOB08(sap)
            # res = ob08.run(rows_data)
            # Lấy export_path từ param2 (nếu có)
            export_path = getattr(task, 'param2', None) or None
            res = ob08.run(rows_data, export_path=export_path)
            log.info(f"  Kết quả OB08: {res}")

            duration = time.time() - start
            
            if res.get('ok'):
                try:
                    os.remove(filepath)
                    log.info(f"  Đã xóa file: {filepath}")
                except Exception as del_err:
                    log.warning(f"  Không thể xóa file: {del_err}")
                    
                return {
                    "status": "success",
                    "message": res.get('message', f"Đã nhập {len(rows_data)} tỷ giá vào OB08"),
                    "rows": len(rows_data),
                    "duration": round(duration, 2)
                }
            else:
                return {
                    "status": "error",
                    "message": res.get('message', 'Lỗi không xác định từ OB08'),
                    "rows": 0,
                    "duration": round(duration, 2)
                }

    except Exception as e:
        log.error(f"  LỖI: {e}")
        return {
            "status": "error",
            "message": str(e),
            "rows": 0,
            "duration": round(time.time() - start, 2)
        }


def md12_unfix_order(filepath, session=None, task=None):
    """
    MD12 - Bỏ tick "Firmly planned" cho Planned Orders
    
    File CSV/Excel: Cột A chứa danh sách Planned Order numbers
    Dòng 1 là tiêu đề, dữ liệu bắt đầu từ dòng 2
    """
    log.info(f"[MD12] Bắt đầu xử lý: {os.path.basename(filepath)}")
    start = time.time()

    try:
        planned_orders = []
        
        if filepath.lower().endswith('.csv'):
            import csv
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                next(reader, None)  # Bỏ qua dòng 1 (tiêu đề)
                for row in reader:
                    if row and row[0].strip():
                        value = row[0].strip()
                        if value.isdigit():
                            planned_orders.append(value)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):  # Bắt đầu từ dòng 2
                cell = row[0]
                if cell.value:
                    value = str(cell.value).strip()
                    if value.isdigit():
                        planned_orders.append(value)
            wb.close()
        
        if not planned_orders:
            return {"status": "skipped", "message": "Không có dữ liệu trong file", "rows": 0}

        log.info(f"  Đọc được {len(planned_orders)} planned orders")

        if not task or not task.sap_user:
            return {
                "status": "success",
                "message": f"Validated (no SAP user): {len(planned_orders)} orders",
                "rows": len(planned_orders)
            }

        sap_client = task.sap_user.client
        sap_username = task.sap_user.username
        sap_password = task.sap_user.password
        log.info(f"  SAP User: {sap_client}/{sap_username}")

        time.sleep(2)

        with SapGuiClient(
            sap_entry_name='V2Q',
            client_no=sap_client,
            username=sap_username,
            password=sap_password,
        ) as sap:
            
            st = sap.last_login_status
            if sap.is_error(st) or ("name or password is incorrect" in (st.text or "").lower()):
                return {
                    "status": "error",
                    "message": f"SAP Login Error: {st.text}",
                    "rows": 0,
                    "duration": round(time.time() - start, 2)
                }
            
            log.info(f"  Đăng nhập SAP thành công")
            
            md12 = TCodeMD12(sap)
            res = md12.run(planned_orders)
            log.info(f"  Kết quả MD12: {res}")

            duration = time.time() - start
            
            if res.get('ok'):
                try:
                    os.remove(filepath)
                    log.info(f"  Đã xóa file: {filepath}")
                except Exception as del_err:
                    log.warning(f"  Không thể xóa file: {del_err}")
                    
                return {
                    "status": "success",
                    "message": res.get('message', f"Đã xử lý {res.get('processed', 0)} planned orders"),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }
            else:
                return {
                    "status": "error",
                    "message": res.get('message', 'Lỗi không xác định từ MD12'),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }

    except Exception as e:
        log.error(f"  LỖI: {e}")
        return {
            "status": "error",
            "message": str(e),
            "rows": 0,
            "duration": round(time.time() - start, 2)
        }


def mm02_update_vietnam_name(filepath, session=None, task=None):
    """
    MM02 - Cập nhật mô tả tiếng Việt cho Material (Tab ZU05)
    
    File Excel/CSV: 
    - Dòng 1: Tiêu đề (bỏ qua)
    - Cột A: Material code
    - Cột B: Mô tả tiếng Việt (Long Text)
    """
    log.info(f"[MM02] Bắt đầu xử lý: {os.path.basename(filepath)}")
    start = time.time()

    try:
        materials = []
        
        if filepath.lower().endswith('.csv'):
            import csv
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                next(reader, None)  # Bỏ qua dòng 1 (tiêu đề)
                for row in reader:
                    if len(row) >= 2 and row[0].strip() and row[1].strip():
                        materials.append({
                            'material': row[0].strip(),
                            'description': row[1].strip()
                        })
        else:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2):  # Bắt đầu từ dòng 2
                material = row[0].value if len(row) > 0 else None
                description = row[1].value if len(row) > 1 else None
                
                if material and description:
                    materials.append({
                        'material': str(material).strip(),
                        'description': str(description).strip()
                    })
            
            wb.close()
        
        if not materials:
            return {"status": "skipped", "message": "Không có dữ liệu trong file", "rows": 0}

        log.info(f"  Đọc được {len(materials)} materials")

        if not task or not task.sap_user:
            return {
                "status": "success",
                "message": f"Validated (no SAP user): {len(materials)} materials",
                "rows": len(materials)
            }

        sap_client = task.sap_user.client
        sap_username = task.sap_user.username
        sap_password = task.sap_user.password
        log.info(f"  SAP User: {sap_client}/{sap_username}")

        time.sleep(2)

        with SapGuiClient(
            sap_entry_name='V2Q',
            client_no=sap_client,
            username=sap_username,
            password=sap_password,
        ) as sap:
            
            st = sap.last_login_status
            if sap.is_error(st) or ("name or password is incorrect" in (st.text or "").lower()):
                return {
                    "status": "error",
                    "message": f"SAP Login Error: {st.text}",
                    "rows": 0,
                    "duration": round(time.time() - start, 2)
                }
            
            log.info(f"  Đăng nhập SAP thành công")
            
            mm02 = TCodeMM02(sap)
            res = mm02.run(materials)
            log.info(f"  Kết quả MM02: {res}")

            duration = time.time() - start
            
            if res.get('ok'):
                try:
                    os.remove(filepath)
                    log.info(f"  Đã xóa file: {filepath}")
                except Exception as del_err:
                    log.warning(f"  Không thể xóa file: {del_err}")
                    
                return {
                    "status": "success",
                    "message": res.get('message', f"Đã xử lý {res.get('processed', 0)} materials"),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }
            else:
                return {
                    "status": "error",
                    "message": res.get('message', 'Lỗi không xác định từ MM02'),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }

    except Exception as e:
        log.error(f"  LỖI: {e}")
        return {
            "status": "error",
            "message": str(e),
            "rows": 0,
            "duration": round(time.time() - start, 2)
        }


def mm02_update_dv_tinh(filepath, session=None, task=None):
    """
    MM02 - Cập nhật đơn vị tính (Tab ZU07 - Internal comment)
    
    File Excel/CSV: 
    - Dòng 1: Tiêu đề (bỏ qua)
    - Cột A: Material code
    - Cột B: Đơn vị tính
    """
    log.info(f"[MM02-ZU07] Bắt đầu xử lý: {os.path.basename(filepath)}")
    start = time.time()

    try:
        materials = []
        
        if filepath.lower().endswith('.csv'):
            import csv
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                next(reader, None)  # Bỏ qua dòng 1 (tiêu đề)
                for row in reader:
                    if len(row) >= 2 and row[0].strip() and row[1].strip():
                        materials.append({
                            'material': row[0].strip(),
                            'description': row[1].strip()
                        })
        else:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2):  # Bắt đầu từ dòng 2
                material = row[0].value if len(row) > 0 else None
                description = row[1].value if len(row) > 1 else None
                
                if material and description:
                    materials.append({
                        'material': str(material).strip(),
                        'description': str(description).strip()
                    })
            
            wb.close()
        
        if not materials:
            return {"status": "skipped", "message": "Không có dữ liệu trong file", "rows": 0}

        log.info(f"  Đọc được {len(materials)} materials")

        if not task or not task.sap_user:
            return {
                "status": "success",
                "message": f"Validated (no SAP user): {len(materials)} materials",
                "rows": len(materials)
            }

        sap_client = task.sap_user.client
        sap_username = task.sap_user.username
        sap_password = task.sap_user.password
        log.info(f"  SAP User: {sap_client}/{sap_username}")

        time.sleep(2)

        with SapGuiClient(
            sap_entry_name='V2Q',
            client_no=sap_client,
            username=sap_username,
            password=sap_password,
        ) as sap:
            
            st = sap.last_login_status
            if sap.is_error(st) or ("name or password is incorrect" in (st.text or "").lower()):
                return {
                    "status": "error",
                    "message": f"SAP Login Error: {st.text}",
                    "rows": 0,
                    "duration": round(time.time() - start, 2)
                }
            
            log.info(f"  Đăng nhập SAP thành công")
            
            mm02 = TCodeMM02(sap)
            res = mm02.run_zu07(materials)
            log.info(f"  Kết quả MM02-ZU07: {res}")

            duration = time.time() - start
            
            if res.get('ok'):
                try:
                    os.remove(filepath)
                    log.info(f"  Đã xóa file: {filepath}")
                except Exception as del_err:
                    log.warning(f"  Không thể xóa file: {del_err}")
                    
                return {
                    "status": "success",
                    "message": res.get('message', f"Đã xử lý {res.get('processed', 0)} materials"),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }
            else:
                return {
                    "status": "error",
                    "message": res.get('message', 'Lỗi không xác định từ MM02'),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }

    except Exception as e:
        log.error(f"  LỖI: {e}")
        return {
            "status": "error",
            "message": str(e),
            "rows": 0,
            "duration": round(time.time() - start, 2)
        }


def me22_change_po_quantity(filepath, session=None, task=None):
    """
    ME22 - Change PO Quantity
    
    File Excel/CSV:
    - Dòng 1: Tiêu đề (bỏ qua)
    - Cột A: PO Number
    - Cột B: Item Number
    - Cột C: Quantity
    """
    log.info(f"[ME22] Bắt đầu xử lý: {os.path.basename(filepath)}")
    start = time.time()

    try:
        po_items = []
        
        if filepath.lower().endswith('.csv'):
            import csv
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                next(reader, None)  # Bỏ qua dòng 1 (tiêu đề)
                for row in reader:
                    if len(row) >= 3 and row[0].strip():
                        po_items.append({
                            'po_number': row[0].strip(),
                            'item_number': row[1].strip(),
                            'quantity': row[2].strip()
                        })
        else:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2):  # Bắt đầu từ dòng 2
                po_number = row[0].value if len(row) > 0 else None
                item_number = row[1].value if len(row) > 1 else None
                quantity = row[2].value if len(row) > 2 else None
                
                if po_number and item_number and quantity:
                    po_items.append({
                        'po_number': str(po_number).strip(),
                        'item_number': str(item_number).strip(),
                        'quantity': str(quantity).strip()
                    })
            
            wb.close()
        
        if not po_items:
            return {"status": "skipped", "message": "Không có dữ liệu trong file", "rows": 0}

        log.info(f"  Đọc được {len(po_items)} PO items")

        if not task or not task.sap_user:
            return {
                "status": "success",
                "message": f"Validated (no SAP user): {len(po_items)} items",
                "rows": len(po_items)
            }

        sap_client = task.sap_user.client
        sap_username = task.sap_user.username
        sap_password = task.sap_user.password
        log.info(f"  SAP User: {sap_client}/{sap_username}")

        time.sleep(2)

        with SapGuiClient(
            sap_entry_name='V2Q',
            client_no=sap_client,
            username=sap_username,
            password=sap_password,
        ) as sap:
            
            st = sap.last_login_status
            if sap.is_error(st) or ("name or password is incorrect" in (st.text or "").lower()):
                return {
                    "status": "error",
                    "message": f"SAP Login Error: {st.text}",
                    "rows": 0,
                    "duration": round(time.time() - start, 2)
                }
            
            log.info(f"  Đăng nhập SAP thành công")
            
            me22 = TCodeME22(sap)
            res = me22.run(po_items)
            log.info(f"  Kết quả ME22: {res}")

            duration = time.time() - start
            
            if res.get('ok'):
                try:
                    os.remove(filepath)
                    log.info(f"  Đã xóa file: {filepath}")
                except Exception as del_err:
                    log.warning(f"  Không thể xóa file: {del_err}")
                    
                return {
                    "status": "success",
                    "message": res.get('message', f"Đã cập nhật {res.get('processed', 0)} PO items"),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }
            else:
                return {
                    "status": "error",
                    "message": res.get('message', 'Lỗi không xác định từ ME22'),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }

    except Exception as e:
        log.error(f"  LỖI: {e}")
        return {
            "status": "error",
            "message": str(e),
            "rows": 0,
            "duration": round(time.time() - start, 2)
        }


def me12_update_marker(filepath, session=None, task=None):
    """
    ME12 - Update Marker for Info Record
    
    File Excel/CSV:
    - Dòng 1: Tiêu đề (bỏ qua)
    - Cột A: Vendor
    - Cột B: Material
    - Cột C: Marker (có thể trống)
    """
    log.info(f"[ME12] Bắt đầu xử lý: {os.path.basename(filepath)}")
    start = time.time()

    try:
        info_records = []
        
        if filepath.lower().endswith('.csv'):
            import csv
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                next(reader, None)  # Bỏ qua dòng 1 (tiêu đề)
                for row in reader:
                    if len(row) >= 2 and row[0].strip() and row[1].strip():
                        info_records.append({
                            'vendor': row[0].strip(),
                            'material': row[1].strip(),
                            'marker': row[2].strip() if len(row) > 2 else ''
                        })
        else:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2):  # Bắt đầu từ dòng 2
                vendor = row[0].value if len(row) > 0 else None
                material = row[1].value if len(row) > 1 else None
                marker = row[2].value if len(row) > 2 else None
                
                if vendor and material:
                    info_records.append({
                        'vendor': str(vendor).strip(),
                        'material': str(material).strip(),
                        'marker': str(marker).strip() if marker else ''
                    })
            
            wb.close()
        
        if not info_records:
            return {"status": "skipped", "message": "Không có dữ liệu trong file", "rows": 0}

        log.info(f"  Đọc được {len(info_records)} info records")

        if not task or not task.sap_user:
            return {
                "status": "success",
                "message": f"Validated (no SAP user): {len(info_records)} records",
                "rows": len(info_records)
            }

        sap_client = task.sap_user.client
        sap_username = task.sap_user.username
        sap_password = task.sap_user.password
        log.info(f"  SAP User: {sap_client}/{sap_username}")

        time.sleep(2)

        with SapGuiClient(
            sap_entry_name='V2Q',
            client_no=sap_client,
            username=sap_username,
            password=sap_password,
        ) as sap:
            
            st = sap.last_login_status
            if sap.is_error(st) or ("name or password is incorrect" in (st.text or "").lower()):
                return {
                    "status": "error",
                    "message": f"SAP Login Error: {st.text}",
                    "rows": 0,
                    "duration": round(time.time() - start, 2)
                }
            
            log.info(f"  Đăng nhập SAP thành công")
            
            me12 = TCodeME12(sap)
            res = me12.run(info_records)
            log.info(f"  Kết quả ME12: {res}")

            duration = time.time() - start
            
            if res.get('ok'):
                try:
                    os.remove(filepath)
                    log.info(f"  Đã xóa file: {filepath}")
                except Exception as del_err:
                    log.warning(f"  Không thể xóa file: {del_err}")
                    
                return {
                    "status": "success",
                    "message": res.get('message', f"Đã cập nhật {res.get('processed', 0)} info records"),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }
            else:
                return {
                    "status": "error",
                    "message": res.get('message', 'Lỗi không xác định từ ME12'),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }

    except Exception as e:
        log.error(f"  LỖI: {e}")
        return {
            "status": "error",
            "message": str(e),
            "rows": 0,
            "duration": round(time.time() - start, 2)
        }


def me12_update_shipping_instruction(filepath, session=None, task=None):
    """
    ME12 - Update Shipping Instruction (EVERS) for Info Record
    
    File Excel/CSV:
    - Dòng 1: Tiêu đề (bỏ qua)
    - Cột A: Vendor
    - Cột B: Material
    - Cột C: Purch. Org
    - Cột D: Plant
    - Cột E: Shipping Instruction (có thể trống)
    """
    log.info(f"[ME12-EVERS] Bắt đầu xử lý: {os.path.basename(filepath)}")
    start = time.time()

    try:
        info_records = []
        
        if filepath.lower().endswith('.csv'):
            import csv
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                next(reader, None)  # Bỏ qua dòng 1 (tiêu đề)
                for row in reader:
                    if len(row) >= 4 and row[0].strip() and row[1].strip():
                        info_records.append({
                            'vendor': row[0].strip(),
                            'material': row[1].strip(),
                            'purch_org': row[2].strip(),
                            'plant': row[3].strip(),
                            'shipping_instr': row[4].strip() if len(row) > 4 else ''
                        })
        else:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2):  # Bắt đầu từ dòng 2
                vendor = row[0].value if len(row) > 0 else None
                material = row[1].value if len(row) > 1 else None
                purch_org = row[2].value if len(row) > 2 else None
                plant = row[3].value if len(row) > 3 else None
                shipping_instr = row[4].value if len(row) > 4 else None
                
                if vendor and material and purch_org and plant:
                    info_records.append({
                        'vendor': str(vendor).strip(),
                        'material': str(material).strip(),
                        'purch_org': str(purch_org).strip(),
                        'plant': str(plant).strip(),
                        'shipping_instr': str(shipping_instr).strip() if shipping_instr else ''
                    })
            
            wb.close()
        
        if not info_records:
            return {"status": "skipped", "message": "Không có dữ liệu trong file", "rows": 0}

        log.info(f"  Đọc được {len(info_records)} info records")

        if not task or not task.sap_user:
            return {
                "status": "success",
                "message": f"Validated (no SAP user): {len(info_records)} records",
                "rows": len(info_records)
            }

        sap_client = task.sap_user.client
        sap_username = task.sap_user.username
        sap_password = task.sap_user.password
        log.info(f"  SAP User: {sap_client}/{sap_username}")

        time.sleep(2)

        with SapGuiClient(
            sap_entry_name='V2Q',
            client_no=sap_client,
            username=sap_username,
            password=sap_password,
        ) as sap:
            
            st = sap.last_login_status
            if sap.is_error(st) or ("name or password is incorrect" in (st.text or "").lower()):
                return {
                    "status": "error",
                    "message": f"SAP Login Error: {st.text}",
                    "rows": 0,
                    "duration": round(time.time() - start, 2)
                }
            
            log.info(f"  Đăng nhập SAP thành công")
            
            me12 = TCodeME12(sap)
            res = me12.run_shipping_instruction(info_records)
            log.info(f"  Kết quả ME12-EVERS: {res}")

            duration = time.time() - start
            
            if res.get('ok'):
                try:
                    os.remove(filepath)
                    log.info(f"  Đã xóa file: {filepath}")
                except Exception as del_err:
                    log.warning(f"  Không thể xóa file: {del_err}")
                    
                return {
                    "status": "success",
                    "message": res.get('message', f"Đã cập nhật {res.get('processed', 0)} info records"),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }
            else:
                return {
                    "status": "error",
                    "message": res.get('message', 'Lỗi không xác định từ ME12'),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }

    except Exception as e:
        log.error(f"  LỖI: {e}")
        return {
            "status": "error",
            "message": str(e),
            "rows": 0,
            "duration": round(time.time() - start, 2)
        }


def mk01_create_maker(filepath, session=None, task=None):
    """
    MK01 - Create Vendor (Maker)
    
    File Excel/CSV:
    - Dòng 1: Tiêu đề (bỏ qua)
    - Cột A: LIFNR (Maker code) - BẮT BUỘC
    - Cột B: NAME1 (Name 1) - BẮT BUỘC
    - Cột C: NAME2 (Name 2)
    - Cột D: EMNFR (Manufacturer number)
    - Cột E: STREET (Street)
    - Cột F: CITY2 (District)
    - Cột G: CITY1 (City)
    - Cột H: COUNTRY (Country code) - BẮT BUỘC
    - Cột I: STCD3 (Tax Number 3)
    - Cột J: STCD4 (Tax Number 4)
    - Cột K: INCO1 (Incoterms 1)
    - Cột L: INCO2 (Incoterms 2)
    
    Fixed values: EKORG=C00A, KTOKK=V090, SORT1=1, WAERS=USD, WEBRE=True
    """
    log.info(f"[MK01] Bắt đầu xử lý: {os.path.basename(filepath)}")
    start = time.time()

    try:
        vendors = []
        skipped_rows = []  # Lưu các dòng bị skip
        
        if filepath.lower().endswith('.csv'):
            import csv
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                next(reader, None)  # Bỏ qua dòng 1 (tiêu đề)
                for idx, row in enumerate(reader):
                    row_number = idx + 2
                    lifnr = row[0].strip() if len(row) > 0 else ''
                    name1 = row[1].strip() if len(row) > 1 else ''
                    country = row[7].strip() if len(row) > 7 else ''
                    
                    # Kiểm tra các trường bắt buộc
                    missing = []
                    if not lifnr:
                        missing.append('LIFNR')
                    if not name1:
                        missing.append('NAME1')
                    if not country:
                        missing.append('COUNTRY')
                    
                    if missing:
                        skipped_rows.append(row_number)
                        continue
                    
                    vendors.append({
                        'lifnr': lifnr,
                        'name1': name1,
                        'name2': row[2].strip() if len(row) > 2 else '',
                        'emnfr': row[3].strip() if len(row) > 3 else '',
                        'street': row[4].strip() if len(row) > 4 else '',
                        'city2': row[5].strip() if len(row) > 5 else '',
                        'city1': row[6].strip() if len(row) > 6 else '',
                        'country': country,
                        'stcd3': row[8].strip() if len(row) > 8 else '',
                        'stcd4': row[9].strip() if len(row) > 9 else '',
                        'inco1': row[10].strip() if len(row) > 10 else '',
                        'inco2': row[11].strip() if len(row) > 11 else '',
                    })
        else:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            for idx, row in enumerate(ws.iter_rows(min_row=2)):
                row_number = idx + 2
                lifnr = str(row[0].value).strip() if len(row) > 0 and row[0].value else ''
                name1 = str(row[1].value).strip() if len(row) > 1 and row[1].value else ''
                country = str(row[7].value).strip() if len(row) > 7 and row[7].value else ''
                
                # Kiểm tra các trường bắt buộc
                missing = []
                if not lifnr:
                    missing.append('LIFNR')
                if not name1:
                    missing.append('NAME1')
                if not country:
                    missing.append('COUNTRY')
                
                if missing:
                    skipped_rows.append(row_number)
                    continue
                
                vendors.append({
                    'lifnr': lifnr,
                    'name1': name1,
                    'name2': str(row[2].value).strip() if len(row) > 2 and row[2].value else '',
                    'emnfr': str(row[3].value).strip() if len(row) > 3 and row[3].value else '',
                    'street': str(row[4].value).strip() if len(row) > 4 and row[4].value else '',
                    'city2': str(row[5].value).strip() if len(row) > 5 and row[5].value else '',
                    'city1': str(row[6].value).strip() if len(row) > 6 and row[6].value else '',
                    'country': country,
                    'stcd3': str(row[8].value).strip() if len(row) > 8 and row[8].value else '',
                    'stcd4': str(row[9].value).strip() if len(row) > 9 and row[9].value else '',
                    'inco1': str(row[10].value).strip() if len(row) > 10 and row[10].value else '',
                    'inco2': str(row[11].value).strip() if len(row) > 11 and row[11].value else '',
                })
            
            wb.close()
        
        # Kiểm tra: có dòng nào thiếu dữ liệu bắt buộc không
        if skipped_rows:
            return {"status": "error", "message": "Các cột MAKER, NAME1, COUNTRY KEY là bắt buộc. Kiểm tra lại file.", "rows": 0}
        
        # Không có dữ liệu
        if not vendors:
            return {"status": "skipped", "message": "Không có dữ liệu trong file", "rows": 0}

        log.info(f"  Đọc được {len(vendors)} vendors")

        if not task or not task.sap_user:
            return {
                "status": "success",
                "message": f"Validated (no SAP user): {len(vendors)} vendors",
                "rows": len(vendors)
            }

        sap_client = task.sap_user.client
        sap_username = task.sap_user.username
        sap_password = task.sap_user.password
        log.info(f"  SAP User: {sap_client}/{sap_username}")

        time.sleep(2)

        with SapGuiClient(
            sap_entry_name='V2Q',
            client_no=sap_client,
            username=sap_username,
            password=sap_password,
        ) as sap:
            
            st = sap.last_login_status
            if sap.is_error(st) or ("name or password is incorrect" in (st.text or "").lower()):
                return {
                    "status": "error",
                    "message": f"SAP Login Error: {st.text}",
                    "rows": 0,
                    "duration": round(time.time() - start, 2)
                }
            
            log.info(f"  Đăng nhập SAP thành công")
            
            mk01 = TCodeMK01(sap)
            res = mk01.run(vendors)
            log.info(f"  Kết quả MK01: {res}")

            duration = time.time() - start
            
            if res.get('ok'):
                try:
                    os.remove(filepath)
                    log.info(f"  Đã xóa file: {filepath}")
                except Exception as del_err:
                    log.warning(f"  Không thể xóa file: {del_err}")
                    
                return {
                    "status": "success",
                    "message": res.get('message', f"Đã tạo {res.get('processed', 0)} vendors"),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }
            else:
                return {
                    "status": "error",
                    "message": res.get('message', 'Lỗi không xác định từ MK01'),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }

    except Exception as e:
        log.error(f"  LỖI: {e}")
        return {
            "status": "error",
            "message": str(e),
            "rows": 0,
            "duration": round(time.time() - start, 2)
        }


def cs02_upload_set_bom_pmg(filepath, session=None, task=None):
    """
    CS02 - Upload/Set BOM PMG
    
    File Excel/CSV:
    - Dòng 1: Tiêu đề (bỏ qua)
    - Cột A: MATNR (Material number - BOM header)
    - Cột B: WERKS (Plant)
    - Cột C: STLAN (BOM usage, default: 1)
    - Cột D: DATUV (Valid from date, format: dd.mm.yyyy)
    - Cột E: IDNRK (Component material)
    - Cột F: MENGE (Quantity)
    - Cột G: MEINS (Unit)
    - Cột H: POSTP (Item category: L, N, etc.)
    - Cột I: LGORT (Storage location)
    """
    log.info(f"[CS02] Bắt đầu xử lý: {os.path.basename(filepath)}")
    start = time.time()

    try:
        # ===== 1. Đọc file =====
        bom_items = []
        
        if filepath.lower().endswith('.csv'):
            import csv
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                next(reader, None)  # Bỏ qua dòng 1 (tiêu đề)
                for row in reader:
                    matnr = row[0].strip() if len(row) > 0 else ''
                    werks = row[1].strip() if len(row) > 1 else ''
                    idnrk = row[4].strip() if len(row) > 4 else ''
                    
                    # Bắt buộc: MATNR, WERKS, IDNRK
                    if matnr and werks and idnrk:
                        bom_items.append({
                            'matnr': matnr,
                            'werks': werks,
                            'stlan': row[2].strip() if len(row) > 2 else '1',
                            'datuv': row[3].strip() if len(row) > 3 else '',
                            'idnrk': idnrk,
                            'menge': row[5].strip() if len(row) > 5 else '',
                            'meins': row[6].strip() if len(row) > 6 else '',
                            'postp': row[7].strip() if len(row) > 7 else '',
                            'lgort': row[8].strip() if len(row) > 8 else '',
                        })
        else:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2):  # Bắt đầu từ dòng 2
                matnr = str(row[0].value).strip() if len(row) > 0 and row[0].value else ''
                werks = str(row[1].value).strip() if len(row) > 1 and row[1].value else ''
                idnrk = str(row[4].value).strip() if len(row) > 4 and row[4].value else ''
                
                # Bắt buộc: MATNR, WERKS, IDNRK
                if matnr and werks and idnrk:
                    bom_items.append({
                        'matnr': matnr,
                        'werks': werks,
                        'stlan': str(row[2].value).strip() if len(row) > 2 and row[2].value else '1',
                        'datuv': str(row[3].value).strip() if len(row) > 3 and row[3].value else '',
                        'idnrk': idnrk,
                        'menge': str(row[5].value).strip() if len(row) > 5 and row[5].value else '',
                        'meins': str(row[6].value).strip() if len(row) > 6 and row[6].value else '',
                        'postp': str(row[7].value).strip() if len(row) > 7 and row[7].value else '',
                        'lgort': str(row[8].value).strip() if len(row) > 8 and row[8].value else '',
                    })
            
            wb.close()
        
        if not bom_items:
            return {"status": "skipped", "message": "Không có dữ liệu trong file", "rows": 0}

        log.info(f"  Đọc được {len(bom_items)} BOM items")

        # ===== 2. Kiểm tra SAP User =====
        if not task or not task.sap_user:
            return {
                "status": "success",
                "message": f"Validated (no SAP user): {len(bom_items)} items",
                "rows": len(bom_items)
            }

        sap_client = task.sap_user.client
        sap_username = task.sap_user.username
        sap_password = task.sap_user.password
        log.info(f"  SAP User: {sap_client}/{sap_username}")

        time.sleep(2)

        # ===== 3. Xử lý trong SAP =====
        with SapGuiClient(
            sap_entry_name='V2Q',
            client_no=sap_client,
            username=sap_username,
            password=sap_password,
        ) as sap:
            
            st = sap.last_login_status
            if sap.is_error(st) or ("name or password is incorrect" in (st.text or "").lower()):
                return {
                    "status": "error",
                    "message": f"SAP Login Error: {st.text}",
                    "rows": 0,
                    "duration": round(time.time() - start, 2)
                }
            
            log.info(f"  Đăng nhập SAP thành công")
            
            # Chạy CS02
            cs02 = TCodeCS02(sap)
            res = cs02.run(bom_items)
            log.info(f"  Kết quả CS02: {res}")

            duration = time.time() - start
            
            if res.get('ok'):
                # Thành công → Xóa file
                try:
                    os.remove(filepath)
                    log.info(f"  Đã xóa file: {filepath}")
                except Exception as del_err:
                    log.warning(f"  Không thể xóa file: {del_err}")
                    
                return {
                    "status": "success",
                    "message": res.get('message', f"Đã thêm {res.get('processed', 0)} components vào BOM"),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }
            else:
                return {
                    "status": "error",
                    "message": res.get('message', 'Lỗi không xác định từ CS02'),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }

    except Exception as e:
        log.error(f"  LỖI: {e}")
        return {
            "status": "error",
            "message": str(e),
            "rows": 0,
            "duration": round(time.time() - start, 2)
        }





def vl32n_del_inbound(filepath, session=None, task=None):
    """
    VL32N - Delete Inbound Delivery
    
    File Excel/CSV:
    - Dòng 1: Tiêu đề (bỏ qua)
    - Cột A: DELIVERY (Delivery Number) - BẮT BUỘC
    """
    log.info(f"[VL32N] Bắt đầu xử lý: {os.path.basename(filepath)}")
    start = time.time()

    try:
        deliveries = []
        skipped_rows = []
        
        if filepath.lower().endswith('.csv'):
            import csv
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                next(reader, None)  # Bỏ qua dòng 1 (tiêu đề)
                for idx, row in enumerate(reader):
                    row_number = idx + 2
                    delivery = row[0].strip() if len(row) > 0 else ''
                    
                    if not delivery:
                        skipped_rows.append(row_number)
                        continue
                    
                    deliveries.append({'delivery': delivery})
        else:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            for idx, row in enumerate(ws.iter_rows(min_row=2)):
                row_number = idx + 2
                delivery = str(row[0].value).strip() if len(row) > 0 and row[0].value else ''
                
                if not delivery:
                    skipped_rows.append(row_number)
                    continue
                
                deliveries.append({'delivery': delivery})
            
            wb.close()
        
        # Kiểm tra dữ liệu
        if skipped_rows:
            return {"status": "error", "message": "Cột DELIVERY là bắt buộc. Kiểm tra lại file.", "rows": 0}
        
        if not deliveries:
            return {"status": "skipped", "message": "Không có dữ liệu trong file", "rows": 0}

        log.info(f"  Đọc được {len(deliveries)} deliveries")

        if not task or not task.sap_user:
            return {
                "status": "success",
                "message": f"Validated (no SAP user): {len(deliveries)} deliveries",
                "rows": len(deliveries)
            }

        sap_client = task.sap_user.client
        sap_username = task.sap_user.username
        sap_password = task.sap_user.password
        log.info(f"  SAP User: {sap_client}/{sap_username}")

        time.sleep(2)

        with SapGuiClient(
            sap_entry_name='V2Q',
            client_no=sap_client,
            username=sap_username,
            password=sap_password,
        ) as sap:
            
            st = sap.last_login_status
            if sap.is_error(st) or ("name or password is incorrect" in (st.text or "").lower()):
                return {
                    "status": "error",
                    "message": f"SAP Login Error: {st.text}",
                    "rows": 0,
                    "duration": round(time.time() - start, 2)
                }
            
            log.info(f"  Đăng nhập SAP thành công")
            
            vl32n = TCodeVL32N(sap)
            res = vl32n.run(deliveries)
            log.info(f"  Kết quả VL32N: {res}")

            duration = time.time() - start
            
            if res.get('ok'):
                try:
                    os.remove(filepath)
                    log.info(f"  Đã xóa file: {filepath}")
                except Exception as del_err:
                    log.warning(f"  Không thể xóa file: {del_err}")
                    
                return {
                    "status": "success",
                    "message": res.get('message', f"Đã xóa {res.get('processed', 0)} deliveries"),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }
            else:
                return {
                    "status": "error",
                    "message": res.get('message', 'Lỗi không xác định từ VL32N'),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }

    except Exception as e:
        log.error(f"  LỖI: {e}")
        return {
            "status": "error",
            "message": str(e),
            "rows": 0,
            "duration": round(time.time() - start, 2)
        }



def me52_del_pr(filepath, session=None, task=None):
    """
    ME52 - Delete Purchase Requisition Item
    
    File Excel/CSV:
    - Dòng 1: Tiêu đề (bỏ qua)
    - Cột A: BANFN (Purchase Requisition Number) - BẮT BUỘC
    - Cột B: BNFPO (Item Number) - BẮT BUỘC
    """
    log.info(f"[ME52] Bắt đầu xử lý: {os.path.basename(filepath)}")
    start = time.time()

    try:
        pr_items = []
        skipped_rows = []
        
        if filepath.lower().endswith('.csv'):
            import csv
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                next(reader, None)  # Bỏ qua dòng 1 (tiêu đề)
                for idx, row in enumerate(reader):
                    row_number = idx + 2
                    banfn = row[0].strip() if len(row) > 0 else ''
                    bnfpo = row[1].strip() if len(row) > 1 else ''
                    
                    # Kiểm tra các trường bắt buộc
                    if not banfn or not bnfpo:
                        skipped_rows.append(row_number)
                        continue
                    
                    pr_items.append({
                        'banfn': banfn,
                        'bnfpo': bnfpo,
                    })
        else:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            for idx, row in enumerate(ws.iter_rows(min_row=2)):
                row_number = idx + 2
                banfn = str(row[0].value).strip() if len(row) > 0 and row[0].value else ''
                bnfpo = str(row[1].value).strip() if len(row) > 1 and row[1].value else ''
                
                # Kiểm tra các trường bắt buộc
                if not banfn or not bnfpo:
                    skipped_rows.append(row_number)
                    continue
                
                pr_items.append({
                    'banfn': banfn,
                    'bnfpo': bnfpo,
                })
            
            wb.close()
        
        # Kiểm tra: có dòng nào thiếu dữ liệu bắt buộc không
        if skipped_rows:
            return {"status": "error", "message": "Các cột BANFN, BNFPO là bắt buộc. Kiểm tra lại file.", "rows": 0}
        
        # Không có dữ liệu
        if not pr_items:
            return {"status": "skipped", "message": "Không có dữ liệu trong file", "rows": 0}

        log.info(f"  Đọc được {len(pr_items)} PR items")

        if not task or not task.sap_user:
            return {
                "status": "success",
                "message": f"Validated (no SAP user): {len(pr_items)} PR items",
                "rows": len(pr_items)
            }

        sap_client = task.sap_user.client
        sap_username = task.sap_user.username
        sap_password = task.sap_user.password
        log.info(f"  SAP User: {sap_client}/{sap_username}")

        time.sleep(2)

        with SapGuiClient(
            sap_entry_name='V2Q',
            client_no=sap_client,
            username=sap_username,
            password=sap_password,
        ) as sap:
            
            st = sap.last_login_status
            if sap.is_error(st) or ("name or password is incorrect" in (st.text or "").lower()):
                return {
                    "status": "error",
                    "message": f"SAP Login Error: {st.text}",
                    "rows": 0,
                    "duration": round(time.time() - start, 2)
                }
            
            log.info(f"  Đăng nhập SAP thành công")
            
            me52 = TCodeME52(sap)
            res = me52.run(pr_items)
            log.info(f"  Kết quả ME52: {res}")

            duration = time.time() - start
            
            if res.get('ok'):
                try:
                    os.remove(filepath)
                    log.info(f"  Đã xóa file: {filepath}")
                except Exception as del_err:
                    log.warning(f"  Không thể xóa file: {del_err}")
                    
                return {
                    "status": "success",
                    "message": res.get('message', f"Đã xóa {res.get('processed', 0)} PR items"),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }
            else:
                return {
                    "status": "error",
                    "message": res.get('message', 'Lỗi không xác định từ ME52'),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }

    except Exception as e:
        log.error(f"  LỖI: {e}")
        return {
            "status": "error",
            "message": str(e),
            "rows": 0,
            "duration": round(time.time() - start, 2)
        }



def mm02_update_ext_matl_group(filepath, session=None, task=None):
    """
    MM02 - Update External Material Group
    
    File Excel/CSV:
    - Dòng 1: Tiêu đề (bỏ qua)
    - Cột A: MATNR (Material Number) - BẮT BUỘC
    - Cột B: EXTWG (External Material Group) - BẮT BUỘC
    """
    log.info(f"[MM02-EXTWG] Bắt đầu xử lý: {os.path.basename(filepath)}")
    start = time.time()

    try:
        materials = []
        skipped_rows = []
        
        if filepath.lower().endswith('.csv'):
            import csv
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                next(reader, None)  # Bỏ qua dòng 1 (tiêu đề)
                for idx, row in enumerate(reader):
                    row_number = idx + 2
                    matnr = row[0].strip() if len(row) > 0 else ''
                    extwg = row[1].strip() if len(row) > 1 else ''
                    
                    if not matnr or not extwg:
                        skipped_rows.append(row_number)
                        continue
                    
                    materials.append({
                        'matnr': matnr,
                        'extwg': extwg,
                    })
        else:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            for idx, row in enumerate(ws.iter_rows(min_row=2)):
                row_number = idx + 2
                matnr = str(row[0].value).strip() if len(row) > 0 and row[0].value else ''
                extwg = str(row[1].value).strip() if len(row) > 1 and row[1].value else ''
                
                if not matnr or not extwg:
                    skipped_rows.append(row_number)
                    continue
                
                materials.append({
                    'matnr': matnr,
                    'extwg': extwg,
                })
            
            wb.close()
        
        # Kiểm tra dữ liệu bắt buộc
        if skipped_rows:
            return {"status": "error", "message": "Các cột A,B là bắt buộc. Kiểm tra lại file.", "rows": 0}
        
        if not materials:
            return {"status": "skipped", "message": "Không có dữ liệu trong file", "rows": 0}

        log.info(f"  Đọc được {len(materials)} materials")

        if not task or not task.sap_user:
            return {
                "status": "success",
                "message": f"Validated (no SAP user): {len(materials)} materials",
                "rows": len(materials)
            }

        sap_client = task.sap_user.client
        sap_username = task.sap_user.username
        sap_password = task.sap_user.password
        log.info(f"  SAP User: {sap_client}/{sap_username}")

        time.sleep(2)

        with SapGuiClient(
            sap_entry_name='V2Q',
            client_no=sap_client,
            username=sap_username,
            password=sap_password,
        ) as sap:
            
            st = sap.last_login_status
            if sap.is_error(st) or ("name or password is incorrect" in (st.text or "").lower()):
                return {
                    "status": "error",
                    "message": f"SAP Login Error: {st.text}",
                    "rows": 0,
                    "duration": round(time.time() - start, 2)
                }
            
            log.info(f"  Đăng nhập SAP thành công")
            
            mm02 = TCodeMM02(sap)
            res = mm02.run_extwg(materials)
            log.info(f"  Kết quả MM02-EXTWG: {res}")

            duration = time.time() - start
            
            if res.get('ok'):
                try:
                    os.remove(filepath)
                    log.info(f"  Đã xóa file: {filepath}")
                except Exception as del_err:
                    log.warning(f"  Không thể xóa file: {del_err}")
                    
                return {
                    "status": "success",
                    "message": res.get('message', f"Đã cập nhật {res.get('processed', 0)} materials"),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }
            else:
                return {
                    "status": "error",
                    "message": res.get('message', 'Lỗi không xác định từ MM02'),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }

    except Exception as e:
        log.error(f"  LỖI: {e}")
        return {
            "status": "error",
            "message": str(e),
            "rows": 0,
            "duration": round(time.time() - start, 2)
        }



def me22_change_price(filepath, session=None, task=None):
    """
    ME22 - Change PO Net Price
    
    File Excel/CSV:
    - Dòng 1: Tiêu đề (bỏ qua)
    - Cột A: PO_NUMBER (Purchase Order Number) - BẮT BUỘC
    - Cột B: ITEM_NUMBER (Item Number) - BẮT BUỘC
    - Cột C: PRICE (Net Price) - BẮT BUỘC
    """
    log.info(f"[ME22-Price] Bắt đầu xử lý: {os.path.basename(filepath)}")
    start = time.time()

    try:
        po_items = []
        skipped_rows = []
        
        if filepath.lower().endswith('.csv'):
            import csv
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                next(reader, None)  # Bỏ qua dòng 1 (tiêu đề)
                for idx, row in enumerate(reader):
                    row_number = idx + 2
                    po_number = row[0].strip() if len(row) > 0 else ''
                    item_number = row[1].strip() if len(row) > 1 else ''
                    price = row[2].strip() if len(row) > 2 else ''
                    
                    if not po_number or not item_number or not price:
                        skipped_rows.append(row_number)
                        continue
                    
                    po_items.append({
                        'po_number': po_number,
                        'item_number': item_number,
                        'price': price,
                    })
        else:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            for idx, row in enumerate(ws.iter_rows(min_row=2)):
                row_number = idx + 2
                po_number = str(row[0].value).strip() if len(row) > 0 and row[0].value else ''
                item_number = str(row[1].value).strip() if len(row) > 1 and row[1].value else ''
                price = str(row[2].value).strip() if len(row) > 2 and row[2].value else ''
                
                if not po_number or not item_number or not price:
                    skipped_rows.append(row_number)
                    continue
                
                po_items.append({
                    'po_number': po_number,
                    'item_number': item_number,
                    'price': price,
                })
            
            wb.close()
        
        # Kiểm tra dữ liệu bắt buộc
        if skipped_rows:
            return {"status": "error", "message": "Các cột A,B,C là bắt buộc. Kiểm tra lại file.", "rows": 0}
        
        if not po_items:
            return {"status": "skipped", "message": "Không có dữ liệu trong file", "rows": 0}

        log.info(f"  Đọc được {len(po_items)} PO items")

        if not task or not task.sap_user:
            return {
                "status": "success",
                "message": f"Validated (no SAP user): {len(po_items)} PO items",
                "rows": len(po_items)
            }

        sap_client = task.sap_user.client
        sap_username = task.sap_user.username
        sap_password = task.sap_user.password
        log.info(f"  SAP User: {sap_client}/{sap_username}")

        time.sleep(2)

        with SapGuiClient(
            sap_entry_name='V2Q',
            client_no=sap_client,
            username=sap_username,
            password=sap_password,
        ) as sap:
            
            st = sap.last_login_status
            if sap.is_error(st) or ("name or password is incorrect" in (st.text or "").lower()):
                return {
                    "status": "error",
                    "message": f"SAP Login Error: {st.text}",
                    "rows": 0,
                    "duration": round(time.time() - start, 2)
                }
            
            log.info(f"  Đăng nhập SAP thành công")
            
            me22 = TCodeME22(sap)
            res = me22.run_price(po_items)
            log.info(f"  Kết quả ME22-Price: {res}")

            duration = time.time() - start
            
            if res.get('ok'):
                try:
                    os.remove(filepath)
                    log.info(f"  Đã xóa file: {filepath}")
                except Exception as del_err:
                    log.warning(f"  Không thể xóa file: {del_err}")
                    
                return {
                    "status": "success",
                    "message": res.get('message', f"Đã cập nhật giá cho {res.get('processed', 0)} PO items"),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }
            else:
                return {
                    "status": "error",
                    "message": res.get('message', 'Lỗi không xác định từ ME22'),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }

    except Exception as e:
        log.error(f"  LỖI: {e}")
        return {
            "status": "error",
            "message": str(e),
            "rows": 0,
            "duration": round(time.time() - start, 2)
        }



def qa11_ud_code(filepath, session=None, task=None):
    """
    QA11 - Record Usage Decision
    
    File Excel/CSV:
    - Dòng 1: Tiêu đề (bỏ qua)
    - Cột A: PRUEFLOS (Inspection Lot Number) - BẮT BUỘC
    - Cột B: VCODE (Usage Decision Code) - BẮT BUỘC
    - Cột C: VCODEGRP (Code Group) - BẮT BUỘC
    """
    log.info(f"[QA11] Bắt đầu xử lý: {os.path.basename(filepath)}")
    start = time.time()

    try:
        items = []
        skipped_rows = []
        
        if filepath.lower().endswith('.csv'):
            import csv
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                next(reader, None)  # Bỏ qua dòng 1 (tiêu đề)
                for idx, row in enumerate(reader):
                    row_number = idx + 2
                    prueflos = row[0].strip() if len(row) > 0 else ''
                    vcode = row[1].strip() if len(row) > 1 else ''
                    vcodegrp = row[2].strip() if len(row) > 2 else ''
                    
                    if not prueflos or not vcode or not vcodegrp:
                        skipped_rows.append(row_number)
                        continue
                    
                    items.append({
                        'prueflos': prueflos,
                        'vcode': vcode,
                        'vcodegrp': vcodegrp,
                    })
        else:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            for idx, row in enumerate(ws.iter_rows(min_row=2)):
                row_number = idx + 2
                prueflos = str(row[0].value).strip() if len(row) > 0 and row[0].value else ''
                vcode = str(row[1].value).strip() if len(row) > 1 and row[1].value else ''
                vcodegrp = str(row[2].value).strip() if len(row) > 2 and row[2].value else ''
                
                if not prueflos or not vcode or not vcodegrp:
                    skipped_rows.append(row_number)
                    continue
                
                items.append({
                    'prueflos': prueflos,
                    'vcode': vcode,
                    'vcodegrp': vcodegrp,
                })
            
            wb.close()
        
        # Kiểm tra dữ liệu bắt buộc
        if skipped_rows:
            return {"status": "error", "message": "Các cột A,B,C là bắt buộc. Kiểm tra lại file.", "rows": 0}
        
        if not items:
            return {"status": "skipped", "message": "Không có dữ liệu trong file", "rows": 0}

        log.info(f"  Đọc được {len(items)} inspection lots")

        if not task or not task.sap_user:
            return {
                "status": "success",
                "message": f"Validated (no SAP user): {len(items)} inspection lots",
                "rows": len(items)
            }

        sap_client = task.sap_user.client
        sap_username = task.sap_user.username
        sap_password = task.sap_user.password
        log.info(f"  SAP User: {sap_client}/{sap_username}")

        time.sleep(2)

        with SapGuiClient(
            sap_entry_name='V2Q',
            client_no=sap_client,
            username=sap_username,
            password=sap_password,
        ) as sap:
            
            st = sap.last_login_status
            if sap.is_error(st) or ("name or password is incorrect" in (st.text or "").lower()):
                return {
                    "status": "error",
                    "message": f"SAP Login Error: {st.text}",
                    "rows": 0,
                    "duration": round(time.time() - start, 2)
                }
            
            log.info(f"  Đăng nhập SAP thành công")
            
            qa11 = TCodeQA11(sap)
            res = qa11.run(items)
            log.info(f"  Kết quả QA11: {res}")

            duration = time.time() - start
            
            if res.get('ok'):
                try:
                    os.remove(filepath)
                    log.info(f"  Đã xóa file: {filepath}")
                except Exception as del_err:
                    log.warning(f"  Không thể xóa file: {del_err}")
                    
                return {
                    "status": "success",
                    "message": res.get('message', f"Đã nhập UD cho {res.get('processed', 0)} inspection lots"),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }
            else:
                return {
                    "status": "error",
                    "message": res.get('message', 'Lỗi không xác định từ QA11'),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }

    except Exception as e:
        log.error(f"  LỖI: {e}")
        return {
            "status": "error",
            "message": str(e),
            "rows": 0,
            "duration": round(time.time() - start, 2)
        }




def vl32n_change_delivery_date(filepath, session=None, task=None):
    """
    VL32N - Change Delivery Date
    
    File Excel/CSV:
    - Dòng 1: Tiêu đề (bỏ qua)
    - Cột A: VBELN (Delivery Number) - BẮT BUỘC
    - Cột B: LFDAT_LA (Delivery Date, dd.mm.yyyy) - BẮT BUỘC
    """
    log.info(f"[VL32N-Date] Bắt đầu xử lý: {os.path.basename(filepath)}")
    start = time.time()

    try:
        items = []
        skipped_rows = []
        
        if filepath.lower().endswith('.csv'):
            import csv
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                next(reader, None)  # Bỏ qua dòng 1 (tiêu đề)
                for idx, row in enumerate(reader):
                    row_number = idx + 2
                    vbeln = row[0].strip() if len(row) > 0 else ''
                    lfdat_la = row[1].strip() if len(row) > 1 else ''
                    
                    if not vbeln or not lfdat_la:
                        skipped_rows.append(row_number)
                        continue
                    
                    items.append({
                        'vbeln': vbeln,
                        'lfdat_la': lfdat_la,
                    })
        else:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            for idx, row in enumerate(ws.iter_rows(min_row=2)):
                row_number = idx + 2
                vbeln = str(row[0].value).strip() if len(row) > 0 and row[0].value else ''
                
                # Xử lý date - có thể là datetime object hoặc string
                lfdat_la = ''
                if len(row) > 1 and row[1].value:
                    cell_value = row[1].value
                    if hasattr(cell_value, 'strftime'):
                        # Là datetime object
                        lfdat_la = cell_value.strftime('%d.%m.%Y')
                    else:
                        lfdat_la = str(cell_value).strip()
                
                if not vbeln or not lfdat_la:
                    skipped_rows.append(row_number)
                    continue
                
                items.append({
                    'vbeln': vbeln,
                    'lfdat_la': lfdat_la,
                })
            
            wb.close()
        
        # Kiểm tra dữ liệu bắt buộc
        if skipped_rows:
            return {"status": "error", "message": "Các cột A,B là bắt buộc. Kiểm tra lại file.", "rows": 0}
        
        if not items:
            return {"status": "skipped", "message": "Không có dữ liệu trong file", "rows": 0}

        log.info(f"  Đọc được {len(items)} deliveries")

        if not task or not task.sap_user:
            return {
                "status": "success",
                "message": f"Validated (no SAP user): {len(items)} deliveries",
                "rows": len(items)
            }

        sap_client = task.sap_user.client
        sap_username = task.sap_user.username
        sap_password = task.sap_user.password
        log.info(f"  SAP User: {sap_client}/{sap_username}")

        time.sleep(2)

        with SapGuiClient(
            sap_entry_name='V2Q',
            client_no=sap_client,
            username=sap_username,
            password=sap_password,
        ) as sap:
            
            st = sap.last_login_status
            if sap.is_error(st) or ("name or password is incorrect" in (st.text or "").lower()):
                return {
                    "status": "error",
                    "message": f"SAP Login Error: {st.text}",
                    "rows": 0,
                    "duration": round(time.time() - start, 2)
                }
            
            log.info(f"  Đăng nhập SAP thành công")
            
            vl32n = TCodeVL32N(sap)
            res = vl32n.run_change_date(items)
            log.info(f"  Kết quả VL32N-Date: {res}")

            duration = time.time() - start
            
            if res.get('ok'):
                try:
                    os.remove(filepath)
                    log.info(f"  Đã xóa file: {filepath}")
                except Exception as del_err:
                    log.warning(f"  Không thể xóa file: {del_err}")
                    
                return {
                    "status": "success",
                    "message": res.get('message', f"Đã cập nhật ngày cho {res.get('processed', 0)} deliveries"),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }
            else:
                return {
                    "status": "error",
                    "message": res.get('message', 'Lỗi không xác định từ VL32N'),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }

    except Exception as e:
        log.error(f"  LỖI: {e}")
        return {
            "status": "error",
            "message": str(e),
            "rows": 0,
            "duration": round(time.time() - start, 2)
        }


def mm02_update_maximum_lot_size(filepath, session=None, task=None):
    """
    MM02 - Update Maximum Lot Size (MRP 1)
    
    File Excel/CSV:
    - Dòng 1: Tiêu đề (bỏ qua)
    - Cột A: MATNR (Material Number) - BẮT BUỘC
    - Cột B: WERKS (Plant) - BẮT BUỘC
    - Cột C: BSTMA (Maximum Lot Size) - BẮT BUỘC
    """
    log.info(f"[MM02-BSTMA] Bắt đầu xử lý: {os.path.basename(filepath)}")
    start = time.time()

    try:
        materials = []
        skipped_rows = []
        
        if filepath.lower().endswith('.csv'):
            import csv
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                next(reader, None)  # Bỏ qua dòng 1 (tiêu đề)
                for idx, row in enumerate(reader):
                    row_number = idx + 2
                    matnr = row[0].strip() if len(row) > 0 else ''
                    werks = row[1].strip() if len(row) > 1 else ''
                    bstma = row[2].strip() if len(row) > 2 else ''
                    
                    if not matnr or not werks or not bstma:
                        skipped_rows.append(row_number)
                        continue
                    
                    materials.append({
                        'matnr': matnr,
                        'werks': werks,
                        'bstma': bstma,
                    })
        else:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            for idx, row in enumerate(ws.iter_rows(min_row=2)):
                row_number = idx + 2
                matnr = str(row[0].value).strip() if len(row) > 0 and row[0].value else ''
                werks = str(row[1].value).strip() if len(row) > 1 and row[1].value else ''
                bstma = str(row[2].value).strip() if len(row) > 2 and row[2].value else ''
                
                if not matnr or not werks or not bstma:
                    skipped_rows.append(row_number)
                    continue
                
                materials.append({
                    'matnr': matnr,
                    'werks': werks,
                    'bstma': bstma,
                })
            
            wb.close()
        
        # Kiểm tra dữ liệu bắt buộc
        if skipped_rows:
            return {"status": "error", "message": "Các cột A,B,C là bắt buộc. Kiểm tra lại file.", "rows": 0}
        
        if not materials:
            return {"status": "skipped", "message": "Không có dữ liệu trong file", "rows": 0}

        log.info(f"  Đọc được {len(materials)} materials")

        if not task or not task.sap_user:
            return {
                "status": "success",
                "message": f"Validated (no SAP user): {len(materials)} materials",
                "rows": len(materials)
            }

        sap_client = task.sap_user.client
        sap_username = task.sap_user.username
        sap_password = task.sap_user.password
        log.info(f"  SAP User: {sap_client}/{sap_username}")

        time.sleep(2)

        with SapGuiClient(
            sap_entry_name='V2Q',
            client_no=sap_client,
            username=sap_username,
            password=sap_password,
        ) as sap:
            
            st = sap.last_login_status
            if sap.is_error(st) or ("name or password is incorrect" in (st.text or "").lower()):
                return {
                    "status": "error",
                    "message": f"SAP Login Error: {st.text}",
                    "rows": 0,
                    "duration": round(time.time() - start, 2)
                }
            
            log.info(f"  Đăng nhập SAP thành công")
            
            mm02 = TCodeMM02(sap)
            res = mm02.run_bstma(materials)
            log.info(f"  Kết quả MM02-BSTMA: {res}")

            duration = time.time() - start
            
            if res.get('ok'):
                try:
                    os.remove(filepath)
                    log.info(f"  Đã xóa file: {filepath}")
                except Exception as del_err:
                    log.warning(f"  Không thể xóa file: {del_err}")
                    
                return {
                    "status": "success",
                    "message": res.get('message', f"Đã cập nhật {res.get('processed', 0)} materials"),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }
            else:
                return {
                    "status": "error",
                    "message": res.get('message', 'Lỗi không xác định từ MM02'),
                    "rows": res.get('processed', 0),
                    "duration": round(duration, 2)
                }

    except Exception as e:
        log.error(f"  LỖI: {e}")
        return {
            "status": "error",
            "message": str(e),
            "rows": 0,
            "duration": round(time.time() - start, 2)
        }